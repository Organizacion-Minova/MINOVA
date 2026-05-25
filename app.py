from flask import Flask, render_template, request, send_file, redirect, url_for
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from fpdf import FPDF
import io

app = Flask(__name__)

# Almacenamiento en memoria (para demo)
datos = []
contador = {"id": 0}


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", datos=datos)


@app.route("/agregar", methods=["POST"])
def agregar():
    nombre    = request.form.get("nombre", "").strip()
    categoria = request.form.get("categoria", "").strip()
    cantidad  = request.form.get("cantidad", "").strip()
    precio    = request.form.get("precio", "").strip()

    if nombre and categoria and cantidad and precio:
        contador["id"] += 1
        datos.append({
            "id":       contador["id"],
            "nombre":   nombre,
            "categoria": categoria,
            "cantidad": int(cantidad),
            "precio":   float(precio),
        })
    return redirect(url_for("index"))


@app.route("/eliminar/<int:fila_id>", methods=["POST"])
def eliminar(fila_id):
    global datos
    datos = [d for d in datos if d["id"] != fila_id]
    return redirect(url_for("index"))


# ── EXPORTAR A EXCEL con openpyxl (con diseño) ───────────────────────────────
@app.route("/exportar/excel")
def exportar_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # ── Estilos reutilizables ────────────────────────────────────────────────
    borde_fino = Border(
        left=Side(style="thin",   color="BFBFBF"),
        right=Side(style="thin",  color="BFBFBF"),
        top=Side(style="thin",    color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    borde_titulo = Border(
        bottom=Side(style="medium", color="1E3A5F"),
    )

    # ── Fila 1: Título del reporte ───────────────────────────────────────────
    ws.merge_cells("A1:E1")
    titulo = ws["A1"]
    titulo.value = "📦  Reporte de Productos"
    titulo.font      = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    titulo.fill      = PatternFill("solid", fgColor="1E3A5F")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Fila 2: Subtítulo / fecha ────────────────────────────────────────────
    from datetime import date
    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value = f"Generado el {date.today().strftime('%d/%m/%Y')}  |  Total registros: {len(datos)}"
    sub.font      = Font(name="Calibri", size=10, italic=True, color="4A6FA5")
    sub.fill      = PatternFill("solid", fgColor="EBF0FA")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Fila 3: vacía de separación ──────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Fila 4: Encabezados de columna ───────────────────────────────────────
    encabezados = ["#", "Nombre", "Categoría", "Cantidad", "Precio (COP)"]
    for col_idx, texto in enumerate(encabezados, start=1):
        cell = ws.cell(row=4, column=col_idx, value=texto)
        cell.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = borde_fino
    ws.row_dimensions[4].height = 22

    # ── Filas de datos ───────────────────────────────────────────────────────
    color_par   = "F0F7FF"   # azul muy suave
    color_impar = "FFFFFF"   # blanco

    formato_precio = '#,##0.00" COP"'
    formato_cant   = '#,##0'

    total_precio = 0
    total_cant   = 0

    for i, d in enumerate(datos, start=1):
        fila_excel = i + 4          # datos empiezan en fila 5
        color_fondo = color_par if i % 2 == 0 else color_impar
        fill_data   = PatternFill("solid", fgColor=color_fondo)

        valores = [i, d["nombre"], d["categoria"], d["cantidad"], d["precio"]]
        alineaciones = ["center", "left", "left", "center", "right"]

        for col_idx, (val, alin) in enumerate(zip(valores, alineaciones), start=1):
            cell = ws.cell(row=fila_excel, column=col_idx, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = fill_data
            cell.alignment = Alignment(horizontal=alin, vertical="center")
            cell.border    = borde_fino

            # Formatos numéricos
            if col_idx == 5:
                cell.number_format = formato_precio
            elif col_idx == 4:
                cell.number_format = formato_cant

        ws.row_dimensions[fila_excel].height = 18
        total_precio += d["precio"]
        total_cant   += d["cantidad"]

    # ── Fila de TOTALES ──────────────────────────────────────────────────────
    if datos:
        fila_total = len(datos) + 5
        ws.row_dimensions[fila_total].height = 22

        etiqueta = ws.cell(row=fila_total, column=1, value="TOTAL")
        etiqueta.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        etiqueta.fill      = PatternFill("solid", fgColor="1E3A5F")
        etiqueta.alignment = Alignment(horizontal="center", vertical="center")
        etiqueta.border    = borde_fino

        # Celdas vacías en columnas 2 y 3
        for col_idx in [2, 3]:
            c = ws.cell(row=fila_total, column=col_idx, value="")
            c.fill   = PatternFill("solid", fgColor="1E3A5F")
            c.border = borde_fino

        # Total cantidad
        cant_cell = ws.cell(row=fila_total, column=4, value=total_cant)
        cant_cell.font          = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cant_cell.fill          = PatternFill("solid", fgColor="1E3A5F")
        cant_cell.alignment     = Alignment(horizontal="center", vertical="center")
        cant_cell.border        = borde_fino
        cant_cell.number_format = formato_cant

        # Total precio
        precio_cell = ws.cell(row=fila_total, column=5, value=total_precio)
        precio_cell.font          = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        precio_cell.fill          = PatternFill("solid", fgColor="1E3A5F")
        precio_cell.alignment     = Alignment(horizontal="right", vertical="center")
        precio_cell.border        = borde_fino
        precio_cell.number_format = formato_precio

    # ── Ancho de columnas ────────────────────────────────────────────────────
    anchos = {"A": 6, "B": 28, "C": 22, "D": 12, "E": 22}
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    # ── Congelar encabezados ─────────────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Guardar y enviar ─────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="tabla_productos.xlsx",
    )


# ── EXPORTAR A PDF con fpdf2 ─────────────────────────────────────────────────
@app.route("/exportar/pdf")
def exportar_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Tabla de Productos", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Encabezados de tabla
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)

    col_w = [10, 55, 45, 30, 40]
    headers = ["#", "Nombre", "Categoría", "Cantidad", "Precio COP"]
    for w, h in zip(col_w, headers):
        pdf.cell(w, 8, h, border=1, fill=True)
    pdf.ln()

    # Filas de datos
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(0, 0, 0)
    for i, d in enumerate(datos, start=1):
        fill = i % 2 == 0
        pdf.set_fill_color(240, 248, 255)
        fila = [str(i), d["nombre"], d["categoria"], str(d["cantidad"]),
                f"${d['precio']:,.0f}"]
        for w, val in zip(col_w, fila):
            pdf.cell(w, 8, val, border=1, fill=fill)
        pdf.ln()

    buffer = io.BytesIO(pdf.output())
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="tabla_productos.pdf",
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)